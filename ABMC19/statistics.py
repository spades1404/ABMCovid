import matplotlib.pyplot as plt
from openpyxl import load_workbook
import ast
import math
wb = load_workbook("output.xlsx")
ws = wb.active


def listAverager(lol):
    length = len(lol[0])
    x = []
    for i in range(length):
        f = []
        for k in lol:
            f.append(k[i])

        z = 0
        for j in f:
            z+=j

        z /= len(f)
        x.append(z)

    return x

#first lets graph average death

n =[list(i) for i in [column for column in ws.iter_rows(min_col=7, max_col=7,min_row=2, values_only=True)]]
n = [list(i) for i in n]
n = ([i[0] for i in n])
n = [ast.literal_eval(i) for i in n]
n = [n[x:x+5] for x in range(0, len(n),5)] #sorts list into mini lists
#n = [sum(i)/len(i) for i in n]
#n = [listAverager(i) for i in n]

print(n)
#x values
w = list(set([int(i[0]) for i in [column for column in ws.iter_rows(min_col=2, max_col=2,min_row=2, values_only=True)]]))
w.sort()


#f = [(n[i]/w[i]) for i in range(len(n))]

'''
plt.plot(w,n)
plt.ylabel("R0")
plt.xlabel("Number of Agents")
#plt.gca().set_ylim([0,None])
plt.show()


for i in n[0]:
    print(i)
    plt.plot(i)
'''


for i in n:

    plt.subplot(5,5,n.index(i)+1)
    for j in i:
        plt.plot(j)
        plt.ylabel("Current Cases")
        plt.xlabel("Ticks")
        plt.title(str((n.index(i)+1)*50 +50))

plt.subplots_adjust(hspace=1)



plt.show()